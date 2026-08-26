# -*- coding: utf-8 -*-
"""
DCF 估值 → 可编辑 Excel 工作簿导出（借鉴 dcf-model skill 的 Excel 模型规范）。

产物：两个 Sheet
  - DCF        ：输入假设(黄底可改) → WACC 自动计算 → 5~N 年 FCFF 预测(全公式)
                 → 终值(Gordon) → EV → 股权价值 → 每股价值 → 隐含收益
  - Sensitivity：WACC × 永续增长率 5×5 敏感性表（每格完整 DCF 公式，中心格=基准）

关键设计（对齐 dcf-model 工作流）：
  1. 公式优先，禁止硬编码：所有预测/现值/终值/敏感性单元格都是活公式，
     用户改任意黄色输入单元格，全表自动联动重算。
  2. 年中折现约定：现金流折现期 0.5/1.5/…，终值折现期 n-0.5（与 run_dcf 一致）。
  3. 敏感性中心格 = 基准假设 → 输出应等于基准每股价值（自校验）。
  4. 蓝色字体=硬编码输入、黑色=公式、黄色填充=可编辑输入（贴近行业习惯，简化为黄底提示可改）。
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..data_fetcher import CompanyData
from .dcf import DCFAssumptions, DCFResult

# ---------- 样式 ----------
FONT_INPUT = Font(color="0000FF", size=11)          # 蓝色 = 硬编码输入
FONT_FORMULA = Font(color="000000", size=11)        # 黑色 = 公式
FONT_BOLD = Font(bold=True, size=11)
FONT_TITLE = Font(bold=True, size=14)
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")    # 深蓝
FILL_SUBHEAD = PatternFill("solid", fgColor="D9E1F2")   # 浅蓝
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")     # 浅黄 = 可编辑输入
FILL_CENTER = PatternFill("solid", fgColor="BDD7EE")    # 中蓝 = 敏感性中心格
FONT_HEADER_WHITE = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PCT = "0.0%"
NUM2 = "#,##0.00"
NUM0 = "#,##0"
NUMNEG = "#,##0;(#,##0);-"


def _set(ws, coord, value, font=None, fill=None, number_format=None, comment=None, border=False):
    """写单元格并应用样式。"""
    c = ws[coord]
    c.value = value
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if number_format:
        c.number_format = number_format
    if border:
        c.border = BORDER_ALL
    if comment:
        from openpyxl.comments import Comment
        cm = Comment(comment, "DCF Panel")
        cm.width, cm.height = 220, 120
        c.comment = cm
    return c


def _fcff_col(n_years: int) -> list[str]:
    """预测期各年所在 Excel 列（D 起）。"""
    return [get_column_letter(4 + i) for i in range(n_years)]


def build_dcf_workbook(cd: CompanyData, assump: DCFAssumptions, dcf_res: DCFResult,
                       wacc: float | None = None, path: str | None = None):
    """构建 DCF 可编辑 Excel 工作簿。path 为空时返回 BytesIO（供面板下载）。"""
    from ..wacc import calc_wacc
    n = int(assump.forecast_years)
    w = float(wacc) if wacc is not None else float(dcf_res.wacc)
    cols = _fcff_col(n)
    last_col = cols[-1]

    # 基础输入（面板已算好，写入作为初始快照；用户可在 Excel 中修改）
    last_rev = dcf_res.assumptions.get("base_rev")
    if last_rev is None or not (last_rev and last_rev > 0):
        last_rev = cd.last_value("revenue")
    debt = cd.last_value("total_debt")
    if not debt or debt < 0:
        debt = 0.0
    cash = cd.last_value("cash")
    if not cash or cash < 0:
        cash = 0.0
    net_debt = float(debt) - float(cash)
    shares = cd.shares
    if not shares or shares <= 0:
        shares = cd.last_value("shares")
    price = cd.latest_price()
    mcap = cd.market_cap if cd.market_cap and cd.market_cap > 0 else shares * price

    wb = Workbook()

    # ================= Sheet 1: DCF =================
    ws = wb.active
    ws.title = "DCF"
    ws.sheet_view.showGridLines = False

    _set(ws, "A1", f"{cd.name}（{cd.symbol}）DCF 估值模型", font=FONT_TITLE)
    _set(ws, "A2", f"Ticker: {cd.symbol} | 市场: {'A股' if cd.market == 'A' else '美股'} | "
                   f"折现约定: 年中(mid-year) | 生成: {cd.latest_year()} 年报基准",
        font=FONT_FORMULA)

    # ---- 输入区 ----
    _set(ws, "A3", "【输入假设】（黄色单元格可编辑，改后全表自动重算）",
         font=FONT_BOLD, fill=FILL_SUBHEAD)
    inputs = [
        ("首年收入增长率", assump.revenue_growth, PCT),
        ("增长率逐年衰减(百分点)", assump.growth_decline, PCT),
        ("增长率逐年加速(百分点)", assump.accel, PCT),
        ("显式预测期年数", n, "0"),
        ("永续增长率 g", assump.terminal_growth, PCT),
        ("目标营业利润率", assump.operating_margin, PCT),
        ("税率", assump.tax_rate, PCT),
        ("资本开支/收入", assump.capex_pct, PCT),
        ("折旧摊销/收入", assump.da_pct, PCT),
        ("营运资本变动/收入(可负)", assump.nwc_pct, PCT),
        ("Beta", assump.beta, "0.00"),
        ("债务成本 Rd", assump.debt_rate, PCT),
        ("无风险利率 rf", assump.rf if assump.rf else cd.rf, PCT),
        ("股权风险溢价 ERP", assump.erp if assump.erp else cd.erp, PCT),
        ("WACC 覆盖(留空=按下方自动计算)", None, PCT),
        ("基准年收入(第0年)", float(last_rev), NUM0),
        ("净债务(债务-现金)", net_debt, NUMNEG),
        ("总股本(股)", float(shares), NUM0),
        ("当前股价", float(price) if price and price > 0 else 0.0, NUM2),
        ("总市值(用于WACC权重)", float(mcap) if mcap and mcap > 0 else 0.0, NUM0),
    ]
    for i, (label, val, nf) in enumerate(inputs):
        r = 4 + i
        _set(ws, f"A{r}", label, font=FONT_FORMULA)
        _set(ws, f"B{r}", val, font=FONT_INPUT, fill=FILL_INPUT, number_format=nf, border=True,
             comment=f"Source: 面板自动生成基准假设；可在此修改，全表公式自动重算。")
    # WACC 覆盖留空提示（B18）
    _set(ws, "B18", None, font=FONT_INPUT, fill=FILL_INPUT, border=True,
         comment="留空则使用下方 WACC 自动计算结果；填入数值将覆盖（如 0.09）。")

    # ---- WACC 自动计算区 ----
    _set(ws, "A24", "【WACC 自动计算】", font=FONT_BOLD, fill=FILL_SUBHEAD)
    _set(ws, "A25", "Re = rf + β × ERP", font=FONT_FORMULA)
    _set(ws, "C25", "=B16+B14*B17", font=FONT_FORMULA, number_format=PCT, border=True)  # rf B16,beta B14,erp B17
    _set(ws, "A26", "股权权重 E/(E+D)", font=FONT_FORMULA)
    _set(ws, "C26", "=IF(B23>0,B23/(B23+MAX(B20,0)),1)", font=FONT_FORMULA, number_format=PCT, border=True)  # 市值 B23
    _set(ws, "A27", "债务权重 D/(E+D)", font=FONT_FORMULA)
    _set(ws, "C27", "=1-C26", font=FONT_FORMULA, number_format=PCT, border=True)
    _set(ws, "A28", "WACC(税后债务加权)", font=FONT_BOLD)
    _set(ws, "C28", "=IF(B18=\"\",C25*C26+B15*(1-B10)*C27,B18)",
         font=FONT_BOLD, number_format=PCT, border=True, fill=FILL_CENTER)
    # 注：B 列行号映射（inputs 从第 4 行起）：
    #   B4 首年增速 B5 衰减 B6 加速 B7 年数 B8 永续g B9 利润率 B10 税率
    #   B11 capex B12 da B13 nwc B14 beta B15 Rd B16 rf B17 erp B18 WACC覆盖
    #   B19 基准收入 B20 净债务 B21 股本 B22 股价 B23 市值

    # ---- 预测区 ----
    _set(ws, "A30", "【预测期】(年中折现 0.5/1.5/…) · 灰色头部=行标签",
         font=FONT_BOLD, fill=FILL_SUBHEAD)
    _set(ws, "C31", "基准年", font=FONT_HEADER_WHITE, fill=FILL_HEADER)
    for j, col in enumerate(cols):
        _set(ws, f"{col}31", f"第{j + 1}年", font=FONT_HEADER_WHITE, fill=FILL_HEADER)
    rows_plan = [
        ("增长率", "=MAX($B$4-$B$5*{i}+$B$6*{i},$B$8-0.01)", PCT),
        ("收入", "={prev}32*(1+{col}30)", NUM0),
        ("EBIT", "={col}32*$B$9", NUM0),
        ("NOPAT", "={col}33*(1-$B$10)", NUM0),
        ("折旧摊销", "={col}32*$B$12", NUM0),
        ("资本开支", "={col}32*$B$11", NUM0),
        ("营运资本变动", "={col}32*$B$13", NUM0),
        ("FCFF", "={col}34+{col}35-{col}36-{col}37", NUM0),
        ("折现期", None, "0.0"),
        ("折现因子", "=1/(1+$C$28)^{col}39", "0.0000"),
        ("FCFF现值", "={col}38*{col}40", NUM0),
    ]
    _set(ws, "C32", "=B19", font=FONT_FORMULA, number_format=NUM0, border=True)
    for j, col in enumerate(cols):
        i = j  # 0-based 年序号
        for rname, (label, formula, nf) in enumerate(rows_plan):
            r = 32 + rname
            if label == "折现期":
                _set(ws, f"{col}{r}", i + 0.5, font=FONT_FORMULA, number_format="0.0", border=True)
            elif formula is not None:
                if label == "增长率":
                    f = formula.format(i=i)
                elif label == "收入":
                    prev = "C" if j == 0 else cols[j - 1]
                    f = formula.format(prev=prev, col=col)
                else:
                    f = formula.format(col=col)
                _set(ws, f"{col}{r}", f, font=FONT_FORMULA, number_format=nf, border=True)

    # ---- 估值区 ----
    _set(ws, "A44", "【估值摘要】", font=FONT_BOLD, fill=FILL_SUBHEAD)
    val_rows = [
        ("显式期 FCFF 现值合计", f"=SUM({cols[0]}41:{last_col}41)", NUM0),
        ("终值 TV = FCFFₙ×(1+g)/(WACC-g)", f"={last_col}38*(1+$B$8)/($C$28-$B$8)", NUM0),
        ("终值现值(年中 折现 n-0.5)", f"=C45/(1+$C$28)^($B$7-0.5)", NUM0),
        ("企业价值 EV", "=C44+C46", NUM0),
        ("(-) 净债务", "=B20", NUMNEG),
        ("股权价值", "=C47-C48", NUM0),
        ("÷ 总股本(股)", "=B21", NUM0),
        ("每股内在价值", "=C49/C50", NUM2),
        ("当前股价", "=B22", NUM2),
        ("隐含收益(相对现价)", "=IF(C52>0,C51/C52-1,\"\")", "0.0%"),
    ]
    for k, (label, formula, nf) in enumerate(val_rows):
        r = 44 + k
        _set(ws, f"A{r}", label, font=FONT_FORMULA if not label.startswith("每股") else FONT_BOLD)
        _set(ws, f"C{r}", formula, font=FONT_BOLD if label in ("每股内在价值", "企业价值 EV", "股权价值") else FONT_FORMULA,
             number_format=nf, border=True,
             fill=FILL_CENTER if label == "每股内在价值" else None)
    _set(ws, "A54", "注：所有蓝色/公式单元格均为活公式；黄色单元格为可编辑输入，修改后整表自动重算。",
         font=Font(italic=True, size=9, color="808080"))

    # ================= Sheet 2: Sensitivity =================
    ws2 = wb.create_sheet("Sensitivity")
    ws2.sheet_view.showGridLines = False
    _set(ws2, "A1", "敏感性分析：每股内在价值 vs (WACC × 永续增长率)", font=FONT_TITLE)
    _set(ws2, "A2", "中心格(蓝底加粗)=基准假设，应等于 DCF 页「每股内在价值」。每格为完整 DCF 公式。",
         font=Font(size=9, color="808080"))

    waccs = [w + (i - 2) * 0.005 for i in range(5)]
    gs = [max(assump.terminal_growth + (i - 2) * 0.005, 0.0) for i in range(5)]

    _set(ws2, "B4", "WACC \\ g", font=FONT_BOLD, fill=FILL_SUBHEAD, border=True)
    for j, g in enumerate(gs):
        col = get_column_letter(3 + j)  # C..G
        _set(ws2, f"{col}4", g, font=FONT_BOLD, number_format=PCT, fill=FILL_SUBHEAD, border=True)
    for i, w_ in enumerate(waccs):
        r = 5 + i
        _set(ws2, f"B{r}", w_, font=FONT_BOLD, number_format=PCT, fill=FILL_SUBHEAD, border=True)
        for j, g in enumerate(gs):
            col = get_column_letter(3 + j)
            term = f"+DCF!{last_col}38*(1+{col}$4)/($B{r}-{col}$4)/(1+$B{r})^($B$7-0.5)"
            fcff_pvs = "+".join(
                f"DCF!{c}38/(1+$B{r})^{k + 0.5}" for k, c in enumerate(cols)
            )
            f = (f"=IF($B{r}<=C$4,\"\","
                 f"({fcff_pvs}{term}-DCF!B20)/DCF!B21)")
            is_center = (i == 2 and j == 2)
            _set(ws2, f"{col}{r}", f,
                 font=FONT_BOLD if is_center else FONT_FORMULA,
                 number_format=NUM2,
                 border=True,
                 fill=FILL_CENTER if is_center else None)
    for col in "ABCDEFG":
        ws2.column_dimensions[col].width = 13

    # ============ Sheet 3: Sensitivity-Growth（收入增长 × 营业利润率） ============
    # 借鉴 dcf-model 工作流的"驱动因素敏感性"：除 WACC×g 外，增长与利润率
    # 是 DCF 价值的两大业务驱动。每格用该格(增长率, 利润率)完整重算 DCF，
    # 引用 DCF 页其余输入（capex/da/nwc/tax/净债务/股本），改动自动联动。
    ws3 = wb.create_sheet("Sensitivity-Growth")
    ws3.sheet_view.showGridLines = False
    _set(ws3, "A1", "敏感性分析：每股内在价值 vs (首年收入增长率 × 营业利润率)", font=FONT_TITLE)
    _set(ws3, "A2", "中心格(蓝底加粗)=基准假设，应等于 DCF 页「每股内在价值」。其余输入（税率/capex/折旧/营运资本/净债务/股本）引用 DCF 页，改动自动联动。",
         font=Font(size=9, color="808080"))

    g0 = float(assump.revenue_growth)
    m0 = float(assump.operating_margin)
    g_grid = [max(g0 + (i - 2) * 0.03, 0.005) for i in range(5)]
    m_grid = [min(max(m0 + (i - 2) * 0.03, 0.01), 0.85) for i in range(5)]

    _set(ws3, "B4", "增长率 \\ 利润率", font=FONT_BOLD, fill=FILL_SUBHEAD, border=True)
    for j, g in enumerate(g_grid):
        col = get_column_letter(3 + j)  # C..G
        _set(ws3, f"{col}4", g, font=FONT_BOLD, number_format=PCT, fill=FILL_SUBHEAD, border=True)
    for i, m in enumerate(m_grid):
        r = 5 + i
        _set(ws3, f"B{r}", m, font=FONT_BOLD, number_format=PCT, fill=FILL_SUBHEAD, border=True)

    def _growth_seq(g_cell: str) -> list[str]:
        """第1~n年增长率公式序列（引用 DCF 页 B5 衰减/B6 加速/B8 终值率）。"""
        return [f"MAX({g_cell}-$B$5*{k}+$B$6*{k},$B$8-0.01)" for k in range(n)]

    def _gm_cell_formula(g_cell: str, m_cell: str) -> str:
        gs = _growth_seq(g_cell)
        # 收入递推（展开乘积）
        revs = []
        prev = "$B$19"
        for k, g in enumerate(gs):
            prev = f"{prev}*(1+{g})"
            revs.append(prev)
        # 单位 FCFF/收入 = margin*(1-tax) + da - capex - nwc（margin 为该格值）
        unit = f"({m_cell}*(1-$B$10)+$B$12-$B$11-$B$13)"
        terms = []
        for k, rev in enumerate(revs):
            pv = f"{rev}*{unit}/(1+$C$28)^{k + 0.5}"
            terms.append(pv)
        tv = (f"{revs[-1]}*{unit}*(1+$B$8)/($C$28-$B$8)"
              f"/(1+$C$28)^($B$7-0.5)")
        body = "+".join(terms + [tv])
        return f"=IF(OR($B$7<=0,{g_cell}<=$B$8-0.011,{m_cell}<=0.005),\"\",({body}-$B$20)/$B$21)"

    for i, m in enumerate(m_grid):
        r = 5 + i
        m_cell = f"$B{r}"
        for j, g in enumerate(g_grid):
            col = get_column_letter(3 + j)
            g_cell = f"{col}$4"
            is_center = (i == 2 and j == 2)
            _set(ws3, f"{col}{r}", _gm_cell_formula(g_cell, m_cell),
                 font=FONT_BOLD if is_center else FONT_FORMULA,
                 number_format=NUM2, border=True,
                 fill=FILL_CENTER if is_center else None)

    for col in "ABCDEFG":
        ws3.column_dimensions[col].width = 13

    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 13
    ws.column_dimensions["A"].width = 34

    if path:
        wb.save(path)
        return path
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def export_dcf_excel(cd: CompanyData, assump: DCFAssumptions, dcf_res: DCFResult,
                     wacc: float | None = None) -> bytes:
    """返回 DCF 工作簿字节流（供 Streamlit download_button 使用）。"""
    return build_dcf_workbook(cd, assump, dcf_res, wacc=wacc)
